import io
import logging
from typing import List, Tuple
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook
from langchain_text_splitters import RecursiveCharacterTextSplitter
from app.config.settings import get_settings

logger = logging.getLogger(__name__)
settings = get_settings()


class DocumentProcessor:
    """Service for processing and extracting text from various document formats."""

    def __init__(self):
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=settings.chunk_size,
            chunk_overlap=settings.chunk_overlap,
            length_function=len,
            separators=["\n\n", "\n", " ", ""]
        )

    def extract_text(self, file_content: bytes, file_type: str) -> str:
        """
        Extract text from file based on file type.

        Args:
            file_content: File content as bytes
            file_type: File extension (pdf, docx, txt, csv, xlsx)

        Returns:
            Extracted text as string
        """
        file_type = file_type.lower().replace('.', '')

        try:
            if file_type == 'pdf':
                return self._extract_from_pdf(file_content)
            elif file_type in ['docx', 'doc']:
                return self._extract_from_docx(file_content)
            elif file_type == 'txt':
                return self._extract_from_txt(file_content)
            elif file_type == 'csv':
                return self._extract_from_csv(file_content)
            elif file_type in ['xlsx', 'xls']:
                return self._extract_from_xlsx(file_content)
            else:
                raise ValueError(f"Unsupported file type: {file_type}")

        except Exception as e:
            logger.error(f"Failed to extract text from {file_type}: {str(e)}")
            raise

    def _extract_from_pdf(self, file_content: bytes) -> str:
        """Extract text from PDF file."""
        pdf_file = io.BytesIO(file_content)
        pdf_reader = PdfReader(pdf_file)

        text = []
        for page_num, page in enumerate(pdf_reader.pages, 1):
            page_text = page.extract_text()
            if page_text:
                text.append(f"[Page {page_num}]\n{page_text}")

        return "\n\n".join(text)

    def _extract_from_docx(self, file_content: bytes) -> str:
        """Extract text from DOCX file."""
        docx_file = io.BytesIO(file_content)
        doc = Document(docx_file)

        text = []
        for para in doc.paragraphs:
            if para.text.strip():
                text.append(para.text)

        return "\n\n".join(text)

    def _extract_from_txt(self, file_content: bytes) -> str:
        """Extract text from TXT file."""
        try:
            # Try UTF-8 first
            return file_content.decode('utf-8')
        except UnicodeDecodeError:
            # Fallback to latin-1
            return file_content.decode('latin-1')

    def _extract_from_csv(self, file_content: bytes) -> str:
        """Extract text from CSV file."""
        csv_text = self._extract_from_txt(file_content)

        # Convert CSV to readable text format
        lines = csv_text.split('\n')
        formatted_lines = []

        for i, line in enumerate(lines):
            if i == 0:
                formatted_lines.append(f"Headers: {line}")
            elif line.strip():
                formatted_lines.append(f"Row {i}: {line}")

        return "\n".join(formatted_lines)

    def _extract_from_xlsx(self, file_content: bytes) -> str:
        """Extract text from XLSX file."""
        xlsx_file = io.BytesIO(file_content)
        workbook = load_workbook(xlsx_file, read_only=True)

        text = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text.append(f"[Sheet: {sheet_name}]")

            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                row_text = " | ".join(
                    [str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text.append(f"Row {row_idx}: {row_text}")

        return "\n".join(text)

    def chunk_text(self, text: str, metadata: dict = None) -> List[Tuple[str, dict]]:
        """
        Split text into chunks with metadata.

        Args:
            text: Text to chunk
            metadata: Base metadata to include with each chunk

        Returns:
            List of (chunk_text, chunk_metadata) tuples
        """
        if not text.strip():
            return []

        chunks = self.text_splitter.split_text(text)

        chunked_data = []
        for idx, chunk in enumerate(chunks):
            chunk_metadata = metadata.copy() if metadata else {}
            chunk_metadata.update({
                "chunk_index": idx,
                "chunk_total": len(chunks)
            })
            chunked_data.append((chunk, chunk_metadata))

        logger.info(f"Split document into {len(chunks)} chunks")
        return chunked_data

    def process_document(self, file_content: bytes, file_type: str, metadata: dict = None) -> List[Tuple[str, dict]]:
        """
        Orchestrate full document processing: extraction and chunking.

        Args:
            file_content: File content as bytes
            file_type: File extension
            metadata: Metadata to include with chunks

        Returns:
            List of (chunk_text, chunk_metadata) tuples
        """
        # Extract text
        text = self.extract_text(file_content, file_type)

        if not text.strip():
            raise ValueError("No text could be extracted from the document")

        # Chunk text
        chunks = self.chunk_text(text, metadata)

        return chunks


def get_document_processor() -> DocumentProcessor:
    """Get DocumentProcessor instance."""
    return DocumentProcessor()
